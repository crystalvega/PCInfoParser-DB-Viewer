from pathlib import Path
from time import sleep

import xlsxwriter as xlwr
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
    
name_sheet1 = 'Общие характеристики'
name_sheet2 = 'Характеристики дисков'
wb = ""
sheet1 = ""
sheet2 = ""

def Write(all, disk, file, allcolumn, diskcolumn, YMD):
    try:
        fileName = file + ' ' + YMD + '.xlsx'
        global wb
        global sheet1
        global sheet2
        
        Create(allcolumn, diskcolumn, fileName)
        wb, sheet1, sheet2 = init(fileName)
        configuration(all)
        configurationdisk(disk)
        return close(fileName, wb)
    except Exception as e:
        print(e)
        return False

def Create(alls, disks, fileName):
    fileObj = Path(fileName)
    raw = 0

    if fileObj.is_file() == False:
        workbook = xlwr.Workbook(fileName, {'strings_to_numbers': True})
        worksheet = workbook.add_worksheet(name_sheet1)
        worksheet2 = workbook.add_worksheet(name_sheet2)
        bold = workbook.add_format({'bold': True})
        while raw < len(alls):
            worksheet.write(0, raw, alls[raw], bold)
            raw += 1
        raw = 0
        while raw < len(disks):
            worksheet2.write(0, raw, disks[raw], bold)
            raw += 1
        workbook.close()
        
def init(fileName):
    wb = load_workbook('./' + fileName)
    return wb, wb[name_sheet1], wb[name_sheet2]
        
def configurationdisk(disk):
    rowinputdisk = 2
    for i2,di in enumerate(disk):
        for i, d in enumerate(di):
            sheet2.cell(row = rowinputdisk, column =i+1).value = d
        sheet2.cell(row = rowinputdisk, column = i+1).value = str(disk[i2][i].day) + '.' + str(disk[i2][i].month) + '.' + str(disk[i2][i].year) + ' ' + str(disk[i2][i].hour) + ':' + str(disk[i2][i].minute) + ':' + str(disk[i2][i].second)
        rowinputdisk += 1

def configuration(all):
    rowinputall = 2
    for i2,al in enumerate(all):
        for i, a in enumerate(al):
            sheet1.cell(row = rowinputall, column =i+1).value = a
        sheet1.cell(row = rowinputall, column = i+1).value = str(all[i2][i].day) + '.' + str(all[i2][i].month) + '.' + str(all[i2][i].year) + ' ' + str(all[i2][i].hour) + ':' + str(all[i2][i].minute) + ':' + str(all[i2][i].second)
        rowinputall += 1
        
def close(fileName, wb):
    while True:
        try:
            wb.save(filename = fileName)
            return True
        except PermissionError:
            return False